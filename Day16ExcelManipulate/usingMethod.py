import openpyxl

#read Excel file
workbook = openpyxl.load_workbook('OwO.xlsx')

#save Excel file
workbook.save('OwO.xlsx')

#1
#Worksheet manipulate

#check all work sheet
print(workbook.sheetnames)
#output: ['sheetnames1', 'sheetnames2']

#choose specific work sheet
print(workbook['sheetname1'])
#output: <Worksheet "sheetname1">

mxR = sheet.max_row #get maximun rows
mxC = sheet.max_column #maximun cols

#add new worksheet at the end
workbook.create_sheet("sheetnames3")

#2
#single saving lattice manipulate

#choose method 1
sheet = workbook['sheetnames1']
print(sheet['B3'.value]) #get value which at row 3 col B

#choose method 2
print(sheet.cell(row=3, column=2).value) #same as upper

